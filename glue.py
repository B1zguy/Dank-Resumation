# from openpyxl import load_workbook
import openpyxl as pyxl
# from typing import Type
from pprint import pprint
from docxtpl import DocxTemplate

# Tmpo load Excel file
tmp_le_xlsx: str = 'example-resumation.xlsx'
tmp_le_docx: str = 'example-resumation.docx'

# Extracting keywords from Excel file
# Returns keyword-placeholder pairs for finding/replacing
def extracting_keywords(file: str) -> dict:
    # Load spreadsheet and target sheet
    file = pyxl.load_workbook(file).active # Apparently going read-only not good idea https://openpyxl.readthedocs.io/en/latest/optimized.html
    # ^^^ Sort this type hinting later

    # Iterate through sheet, skipping header, then adding to dict
    keywords: dict = {}
    for row in file.iter_rows(min_row=2, values_only=True):
        # print(type(file.iter_rows))
        row: tuple
        # print(row)
        # Extract row's columns to key/val pairs
        keywords[row[0]] = row[1]

    return keywords
# pprint(extracting_keywords(tmp_le_xlsx))
xl = extracting_keywords(tmp_le_xlsx)

# Replacing placeholders in Word file with keywords from Excel
# which is now loaded as a dict to perform find/replace
def replacing_placeholders(file: str, findReplace: dict):
    replacedDoc = file.replace('.docx', ' REPLACED.docx')
    # Loads WOrd file --> Renders doc w/ replaced words --> Saves file
    file = DocxTemplate(file).render(findReplace).save(replacedDoc)
    # ^^^ Sort this type hinting later

    return
replacing_placeholders(tmp_le_docx, xl)
