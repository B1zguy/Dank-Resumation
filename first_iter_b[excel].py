from openpyxl import load_workbook
from pprint import pprint

fp = 'example-resumation.xlsx'
wb = load_workbook(fp) # set read-only in production

sheet = wb.active

# WORKING #
'''a = sheet['A2':'B3']
for i in a:
    print(i[0].value, i[1].value)'''

for row in sheet.iter_rows(min_row=2, values_only=True): # Gives iterable obj. Can specify where begin iterating.
    print(row)                                           # Set to just return values. Useful skip header!
# OR
pprint(dict(sheet.rows)) # grabs all cells at once. Can even return as list() or tuple()
                         # Perhaps create list comprehension to convert all cell obks to vals



# https://openpyxl.readthedocs.io/en/stable/tutorial.html?highlight=ws.iter_rows()#accessing-many-cells